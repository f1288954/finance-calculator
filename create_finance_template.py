from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_finance_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "金融計算模板 Finance Calculator"
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # 設定樣式
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 標題行
    headers = ["計算類別 Calculation Type", "變數說明 Variables", "數值 Value", "Excel公式 Formula", "結果 Result"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # 計算項目
    calculations = [
        ("現值計算 Present Value (PV)", 
         ["未來值 Future Value (FV)", "利率 Interest Rate (r)", "期數 Number of Periods (n)"], 
         "=PV(C3,C4,0,-C2,0)"),
        
        ("未來值計算 Future Value (FV)", 
         ["現值 Present Value (PV)", "利率 Interest Rate (r)", "期數 Number of Periods (n)"], 
         "=FV(C7,C8,0,-C6,0)"),
        
        ("普通年金現值 Present Value of Ordinary Annuity", 
         ["每期金額 Payment (PMT)", "利率 Interest Rate (r)", "期數 Number of Periods (n)"], 
         "=PV(C11,C12,-C10,0,0)"),
        
        ("普通年金終值 Future Value of Ordinary Annuity", 
         ["每期金額 Payment (PMT)", "利率 Interest Rate (r)", "期數 Number of Periods (n)"], 
         "=FV(C15,C16,-C14,0,0)"),
        
        ("成長年金現值 Present Value of Growing Annuity", 
         ["首期金額 First Payment (C1)", "利率 Interest Rate (r)", "成長率 Growth Rate (g)", "期數 Number of Periods (n)"], 
         "=C18/(C19-C20)*(1-POWER(1+C20,C21)/POWER(1+C19,C21))"),
        
        ("永續年金 Perpetuity", 
         ["每期金額 Payment (C)", "利率 Interest Rate (r)"], 
         "=C23/C24"),
        
        ("成長型永續年金 Growing Perpetuity", 
         ["首期金額 First Payment (C1)", "利率 Interest Rate (r)", "成長率 Growth Rate (g)"], 
         "=C26/(C27-C28)"),
        
        ("淨現值計算 Net Present Value (NPV)", 
         ["利率 Interest Rate (r)", "期初投資 Initial Investment (CF0)", 
          "現金流1 Cash Flow 1", "現金流2 Cash Flow 2", "現金流3 Cash Flow 3", 
          "現金流4 Cash Flow 4", "現金流5 Cash Flow 5"], 
         "=NPV(C30,C32:C36)+C31"),
        
        ("內部報酬率 Internal Rate of Return (IRR)", 
         ["期初投資 Initial Investment (CF0)", 
          "現金流1 Cash Flow 1", "現金流2 Cash Flow 2", "現金流3 Cash Flow 3", 
          "現金流4 Cash Flow 4", "現金流5 Cash Flow 5"], 
         "=IRR(C38:C43)"),
        
        ("股利折現模型 Dividend Discount Model (DDM)", 
         ["明年股利 Next Year Dividend (D1)", "必要報酬率 Required Return (r)", "成長率 Growth Rate (g)"], 
         "=C45/(C46-C47)"),
        
        ("本益比 Price-Earnings Ratio (P/E)", 
         ["保留盈餘率 Retention Ratio (b)", "必要報酬率 Required Return (r)", "成長率 Growth Rate (g)"], 
         "=(1-C49)/(C50-C51)"),
        
        ("債券價格 Bond Price", 
         ["票面利率 Coupon Rate (C)", "市場利率 Market Rate (r)", 
          "期數 Number of Periods (n)", "面額 Face Value (F)"], 
         "=PV(C54,C55,-C53*C56,C56,0)"),
        
        ("債券殖利率 Yield to Maturity (YTM)", 
         ["市價 Market Price (P)", "票面利率 Coupon Rate (C)", 
          "期數 Number of Periods (n)", "面額 Face Value (F)"], 
         "=RATE(C59,C58*C60,-C57,C60)")
    ]
    
    current_row = 2
    for calc_type, variables, formula in calculations:
        # 計算類別
        ws.cell(row=current_row, column=1, value=calc_type).border = border
        
        # 變數
        for i, var in enumerate(variables):
            row = current_row + i
            ws.cell(row=row, column=2, value=var).border = border
            ws.cell(row=row, column=3).border = border  # 數值欄位
        
        # 公式和結果
        result_cell = ws.cell(row=current_row, column=4)
        result_cell.value = formula
        result_cell.border = border
        
        ws.cell(row=current_row, column=5).border = border  # 結果欄位
        
        # 更新行數（加上變數數量和一個空行）
        current_row += len(variables) + 1
    
    # 保存文件
    wb.save('金融計算模板.xlsx')

if __name__ == "__main__":
    create_finance_template()
